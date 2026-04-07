import csv
import os
import re 
from glob import glob
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# === Étape 1 : Trouver le dernier fichier CSV clients_*.csv ===
csv_files = glob("clients_*.csv")
if not csv_files:
    print("❌ Aucun fichier clients_*.csv trouvé.")
    exit()

# Trier par date de modification (ou par nom si plus simple)
latest_csv = max(csv_files, key=os.path.getmtime)
print(f"Dernier fichier détecté : {latest_csv}")

# === Étape 2 : Préparer la sortie Excel ===
xlsx_output = latest_csv.replace(".csv", "_multi.xlsx")

# === Étape 3 : Lire les données du CSV ===
with open(latest_csv, newline='') as f:
    reader = csv.DictReader(f)
    all_clients = list(reader)

if not all_clients:
    print("❗ Fichier CSV vide ou mal formaté.")
    exit()

# === Étape 4 : Regrouper les clients par ESSID ===
clients_by_essid = defaultdict(list)
for client in all_clients:
    essid = client.get("essid", "Inconnu")
    clients_by_essid[essid].append(client)

# Nettoyer le nom de la table
def clean_name(name):
    # Remplace espaces et tirets par underscore
    name = re.sub(r"[ -]", "_", name)
    # Supprime tout ce qui n'est pas alphanumérique ou underscore
    name = re.sub(r"[^\w]", "", name)
    # Limite la longueur à 31 caractères (limite Excel)
    return name[:31]


# === Étape 5 : Création du fichier Excel ===
wb = Workbook()
first = True
headers = ["name", "ip", "mac", "os", "essid", "ap", "timestamp"]

for essid, clients in clients_by_essid.items():
    if first:
        ws = wb.active
        ws.title = essid[:31]
        first = False
    else:
        ws = wb.create_sheet(title=essid[:31])

    ws.append(headers)

    for client in clients:
        ws.append([client.get(h, "") for h in headers])

    # Définir un tableau filtrable
    end_col = get_column_letter(len(headers))
    end_row = len(clients) + 1
    table_ref = f"A1:{end_col}{end_row}"

    table = Table(displayName=f"{clean_name(essid)}_Table", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

# === Étape 6 : Sauvegarde du fichier Excel ===
wb.save(xlsx_output)
print(f"Fichier Excel multi-feuilles généré : {xlsx_output}")